import os, csv, io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import letter, portrait
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import inch
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from django.utils import timezone

def create_excel_report(queryset, title, description):
    """
    Create an Excel report with the given queryset, title, and description.
    """
    # Create a new workbook and add a worksheet
    wb = Workbook()
    ws = wb.active
    
    # Clean worksheet title (remove special characters that might cause Excel warnings)
    ws.title = "SitInReport"[:31]  # Excel sheet names max 31 chars
    
    # Add university header (centered and merged across all columns)
    header_rows = [
        ("University of Cebu - Main", 12, True),
        ("College of Computer Studies", 12, True),
        ("Computer Laboratory Sitin Monitoring System Report", 12, True),
        ("", 10, False),  # Empty row for spacing
        (title, 14, True),
        (description, 12, False),
        ("", 10, False),  # Empty row for spacing
    ]

    for i, (text, size, is_bold) in enumerate(header_rows, start=1):
        ws.append([text])
        cell = ws[f'A{i}']
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells(f'A{i}:H{i}')  # Merge across all data columns
        
        # Apply formatting
        cell.font = Font(size=size, bold=is_bold, italic=(i == 6))  # Description is italic
        
        # Add light background to main headers
        if i <= 3:
            cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    # Add column headers
    headers = [
        "Student ID",
        "Full Name",
        "Purpose",
        "Lab Room",
        "Status",
        "Sessions",
        "Time-in",
        "Time-out",
    ]
    ws.append(headers)
    
    # Format header row
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for col in range(1, 9):  # Columns A-H
        cell = ws.cell(row=8, column=col)  # Header is row 8 (after 7 header rows)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Add data rows with proper formatting
    for sitin in queryset:
        fullname = f"{sitin.user.registration.firstname} {sitin.user.registration.lastname}"
        student_id = str(sitin.user.registration.idno)  # Ensure ID is string
        
        # Format datetimes in Excel-friendly way
        time_in = timezone.localtime(sitin.sitin_date).strftime("%b %d, %Y, %I:%M %p") if sitin.sitin_date else ""
        time_out = timezone.localtime(sitin.logout_date).strftime("%b %d, %Y, %I:%M %p") if sitin.logout_date else ""

        ws.append([
            student_id,
            fullname,
            sitin.purpose,
            sitin.lab_room,
            sitin.status,
            sitin.user.registration.sessions if hasattr(sitin.user, "registration") else "",
            time_in,
            time_out,
        ])

    # Set column widths
    column_widths = {
        'A': 15,  # Student ID
        'B': 25,  # Full Name
        'C': 20,  # Purpose
        'D': 15,  # Lab Room
        'E': 15,  # Status
        'F': 15,  # Sessions
        'G': 20,  # Time-in
        'H': 20,  # Time-out
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Freeze header row
    ws.freeze_panes = 'A9'  # Freeze above row 9 (where data starts)

    return wb

def create_csv_report(queryset, title, description):
    if queryset.exists() and title and description:
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Add university headers
        writer.writerow(["University of Cebu - Main"])
        writer.writerow(["College of Computer Studies"])
        writer.writerow(["Computer Laboratory Sitin Monitoring System Report"])
        writer.writerow([])  # Empty row for spacing
        
        # Add report title and description
        writer.writerow([title])
        writer.writerow([description])
        writer.writerow([])  # Empty row for spacing
        
        # Add column headers
        headers = [
            "Student ID",
            "Full Name",
            "Purpose",
            "Lab Room",
            "Status",
            "Sessions",
            "Time-in",
            "Time-out",
        ]
        writer.writerow(headers)
        
        # Add data rows
        for sitin in queryset:
            fullname = f"{sitin.user.registration.firstname} {sitin.user.registration.lastname}"
            student_id = sitin.user.registration.idno

            # Time-zone aware datetimes
            time_in = timezone.localtime(sitin.sitin_date).strftime("%b %d, %Y, %I:%M %p") if sitin.sitin_date else ""
            time_out = timezone.localtime(sitin.logout_date).strftime("%b %d, %Y, %I:%M %p") if sitin.logout_date else ""

            writer.writerow([
                student_id, 
                fullname, 
                sitin.purpose, 
                sitin.lab_room,
                sitin.status, 
                getattr(sitin.user.registration, "sessions", ""),
                time_in, 
                time_out
            ])
        
        return output.getvalue()
    print('create_csv_report error: Missing queryset, title or description')
    return None
    
def create_pdf_report(queryset, title, description):
    buffer = io.BytesIO()
    pdf = SimpleDocTemplate(buffer, pagesize=portrait(letter))
    elements = []
    styles = getSampleStyleSheet()
    
    # Create custom styles
    university_header_style = ParagraphStyle(
        name='UniversityHeader',
        parent=styles['Normal'],
        fontSize=14,
        leading=18,
        alignment=1,
        spaceAfter=6,
        textColor=colors.HexColor('#003366'),
        fontName='Helvetica-Bold'
    )
    
    report_title_style = ParagraphStyle(
        name='ReportTitle',
        parent=styles['Normal'],
        fontSize=16,
        leading=20,
        alignment=1,
        spaceAfter=12,
        textColor=colors.HexColor('#006600'),
        fontName='Helvetica-Bold'
    )
    
    table_header_style = ParagraphStyle(
        name='TableHeader',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.white,
        fontName='Helvetica-Bold',
        leading=12
    )
    
    # Add content
    elements.append(Paragraph("UNIVERSITY OF CEBU - MAIN", university_header_style))
    elements.append(Paragraph("COLLEGE OF COMPUTER STUDIES", university_header_style))
    elements.append(Paragraph("COMPUTER LABORATORY SIT-IN MONITORING SYSTEM", university_header_style))
    elements.append(Spacer(1, 24))
    
    elements.append(Spacer(width=7*inch, height=0.1*inch))
    elements[-1].backColor = colors.HexColor('#003366')
    
    elements.append(Paragraph(title.upper(), report_title_style))
    elements.append(Paragraph(description, styles['Normal']))
    elements.append(Spacer(1, 24))
    
    # Table data
    headers = [
        "Student ID", "Full Name", "Purpose", 
        "Lab Room", "Status", "Sessions", 
        "Time-in", "Time-out"
    ]
    
    # Create header row with white text
    data = [[Paragraph(header, table_header_style) for header in headers]]

    for sitin in queryset:
        fullname = f"{sitin.user.registration.firstname} {sitin.user.registration.lastname}"
        student_id = sitin.user.registration.idno
        sessions = getattr(sitin.user.registration, "sessions", 0)

        time_in = timezone.localtime(sitin.sitin_date).strftime("%b %d, %Y, %I:%M %p") if sitin.sitin_date else ""
        time_out = timezone.localtime(sitin.logout_date).strftime("%b %d, %Y, %I:%M %p") if sitin.logout_date else ""

        data.append([
            Paragraph(str(student_id), styles['Normal']),
            Paragraph(fullname, styles['Normal']),
            Paragraph(str(sitin.purpose), styles['Normal']),
            Paragraph(str(sitin.lab_room), styles['Normal']),
            Paragraph(str(sitin.status), styles['Normal']),
            Paragraph(str(sessions), styles['Normal']),
            Paragraph(time_in, styles['Normal']),
            Paragraph(time_out, styles['Normal']),
        ])
    
    # Create table
    col_widths = [0.7*inch, 1.5*inch, 1.2*inch, 0.8*inch, 0.8*inch, 0.8*inch, 1.2*inch, 1.2*inch]
    table = Table(data, colWidths=col_widths, repeatRows=1)
    
    # Table style - the key is to set text color in both the style and cell content
    style = TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#003366')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 10),
        ('BOTTOMPADDING', (0,0), (-1,0), 8),
        ('BACKGROUND', (0,1), (-1,-1), colors.white),
        ('GRID', (0,0), (-1,-1), 0.5, colors.lightgrey),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F5F5F5')]),
    ])
    table.setStyle(style)
    elements.append(table)
    
    # Footer
    elements.append(Spacer(1, 12))
    elements.append(Paragraph(
        f"Generated on {timezone.localtime(timezone.now()).strftime('%B %d, %Y at %I:%M %p')}",
        ParagraphStyle(name='Footer', fontSize=8, alignment=2)
    ))
    
    pdf.build(elements)
    buffer.seek(0)
    return buffer.getvalue()