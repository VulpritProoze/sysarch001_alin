import os, csv, io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

def create_excel_report(queryset, title, description):
    """
    Create an Excel report with the given queryset, title, and description.
    """
    # Create a new workbook and add a worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Sit-in History"

    # Add title and description
    ws.append([title])
    ws.append([description])
    ws.append([])  # Empty row for spacing

    # Add headers
    headers = [
        "Student ID",
        "Full Name",
        "Purpose",
        "Lab Room",
        "Status",
        "Sessions",
        "Request Date",
        "Time-in",
        "Time-out",
    ]
    ws.append(headers)

    # Add data rows
    for sitin in queryset:
        fullname = sitin.user.registration.firstname + " " + sitin.user.registration.lastname
        student_id = sitin.user.registration.idno

        # Convert timezone-aware datetimes to timezone-naive
        request_date = sitin.date.strftime("%m-%d-%Y %H:%M:%S") if sitin.date else None
        time_in = sitin.sitin_date.strftime("%m-%d-%Y %H:%M:%S") if sitin.sitin_date else None
        time_out = sitin.logout_date.strftime("%m-%d-%Y %H:%M:%S") if sitin.logout_date else None


        ws.append([
            student_id,
            fullname,
            sitin.purpose,
            sitin.lab_room,
            sitin.status,
            sitin.user.registration.sessions if hasattr(sitin.user, "registration") else "",
            request_date,
            time_in,
            time_out,
        ])

    # Apply formatting to the title and description
    title_cell = ws['A1']
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal='center')

    description_cell = ws['A2']
    description_cell.font = Font(size=12, italic=True)

    # Merge cells for the title and description
    ws.merge_cells('A1:I1')  # Merge cells for the title
    ws.merge_cells('A2:I2')  # Merge cells for the description

    # Set column widths
    column_widths = {
        'A': 15,  # Student ID
        'B': 25,  # Full Name
        'C': 20,  # Purpose
        'D': 15,  # Lab Room
        'E': 15,  # Status
        'F': 15,  # Sessions
        'G': 20,  # Request Date
        'H': 20,  # Time-in
        'I': 20,  # Time-out
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    return wb

def create_csv_report(queryset, title, description):
    if queryset.exists() and title and description:
        output = io.StringIO()
        writer = csv.writer(output)
        l = []
        l.append(title)
        writer.writerow(l)
        l.pop()
        l.append(description)
        l.pop()
        writer.writerow(l)
        writer.writerow(l)
        headers = [
            "Student ID",
            "Full Name",
            "Purpose",
            "Lab Room",
            "Status",
            "Sessions",
            "Request Date",
            "Time-in",
            "Time-out",
        ]
        writer.writerow(headers)
        for sitin in queryset:
            fullname = f"{sitin.user.registration.firstname} {sitin.user.registration.lastname}"
            student_id = sitin.user.registration.idno
            # Time-zone naive datetimes
            request_date = sitin.date.strftime("%m-%d-%Y %H:%M:%S") if sitin.date else None
            time_in = sitin.sitin_date.strftime("%m-%d-%Y %H:%M:%S") if sitin.sitin_date else None
            time_out = sitin.logout_date.strftime("%m-%d-%Y %H:%M:%S") if sitin.logout_date else None

            writer.writerow([
                student_id, 
                fullname, 
                sitin.purpose, 
                sitin.lab_room,
                sitin.status, getattr(sitin.user.registration, "sessions", ""),
                request_date, 
                time_in, 
                time_out
            ])
        
        return output.getvalue()
    print('create_csv_report error')
    
def create_pdf_report(queryset, title, description):
    buffer = io.BytesIO()
    pdf = SimpleDocTemplate(buffer, pagesize=landscape(letter))
    elements = []
    styles =  getSampleStyleSheet()
    styleN = styles["Normal"]
    
    # Title
    elements.append(Paragraph(f"<b>{title}</b>", styles['Title']))
    elements.append(Spacer(1, 12))

    # Description
    elements.append(Paragraph(description, styles['Normal']))
    elements.append(Spacer(1,24))
    
    # Content
    headers = [
            "Student ID",
            "Full Name",
            "Purpose",
            "Lab Room",
            "Status",
            "Sessions",
            "Request Date",
            "Time-in",
            "Time-out",
        ]
    data = [[Paragraph(f"<b>{header}</b>", styles['BodyText']) for header in headers]]

    for sitin in queryset:
        fullname = f"{sitin.user.registration.firstname} {sitin.user.registration.lastname}"
        student_id = sitin.user.registration.idno
        sessions = sitin.user.registration.sessions if sitin.user.registration.sessions else 0
        # Time-zone naive datetimes
        request_date = sitin.date.strftime("%m-%d-%Y %H:%M:%S") if sitin.date else None
        time_in = sitin.sitin_date.strftime("%m-%d-%Y %H:%M:%S") if sitin.sitin_date else None
        time_out = sitin.logout_date.strftime("%m-%d-%Y %H:%M:%S") if sitin.logout_date else None

        row = [
            Paragraph(str(student_id), styleN),
            Paragraph(fullname, styleN),
            Paragraph(sitin.purpose, styleN),
            Paragraph(sitin.lab_room, styleN),
            Paragraph(sitin.status, styleN),
            Paragraph(str(sessions), styleN),
            Paragraph(request_date if request_date else "", styleN),
            Paragraph(time_in if time_in else "", styleN),
            Paragraph(time_out if time_out else "", styleN),
        ]
        data.append(row)
    table = Table(data, colWidths=[60, 100, 80, 60, 60, 60, 80, 80, 80])
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),  # Left-align for readability
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),  # Align text to the top
    ])
    table.setStyle(style)
    elements.append(table)
    pdf.build(elements)
    buffer.seek(0)
    return buffer.getvalue()
    